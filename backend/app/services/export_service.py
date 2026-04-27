"""报告导出服务 — Excel"""
import io
import uuid

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.report import TestReport, TestReportScenario


STATUS_LABELS = {"passed": "通过", "failed": "失败", "error": "错误", "skipped": "跳过", "pending": "待录入"}
STATUS_COLORS = {"passed": "6ECF96", "failed": "F08A8E", "error": "F5B87A", "skipped": "BFC4CD", "pending": "A78BFA"}


async def export_excel(
    session: AsyncSession,
    plan_id: uuid.UUID | None = None,
    report_id: uuid.UUID | None = None,
) -> io.BytesIO | None:
    """生成 Excel 报告，返回 BytesIO 文件流。"""
    if report_id:
        result = await session.execute(
            select(TestReport).where(TestReport.id == report_id)
        )
    elif plan_id:
        result = await session.execute(
            select(TestReport).where(TestReport.plan_id == plan_id).order_by(TestReport.created_at.desc())
        )
    else:
        return None
    report = result.scalars().first()
    if report is None:
        return None

    # 获取 scenarios
    scenarios_result = await session.execute(
        select(TestReportScenario)
        .where(TestReportScenario.report_id == report.id)
        .order_by(TestReportScenario.sort_order)
    )
    scenarios = scenarios_result.scalars().all()

    # 创建 Excel
    wb = Workbook()

    # Sheet 1: 汇总
    ws_summary = wb.active
    ws_summary.title = "汇总"
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="6B7EF5", end_color="6B7EF5", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='E8E8EC'),
        right=Side(style='thin', color='E8E8EC'),
        top=Side(style='thin', color='E8E8EC'),
        bottom=Side(style='thin', color='E8E8EC'),
    )

    summary_data = [
        ["指标", "数值"],
        ["总用例数", report.total_scenarios],
        ["通过", report.passed],
        ["失败", report.failed],
        ["错误", report.error],
        ["跳过", report.skipped],
        ["通过率", f"{report.pass_rate}%" if report.pass_rate is not None else "-"],
        ["执行时间", report.executed_at.strftime("%Y-%m-%d %H:%M:%S") if report.executed_at else "-"],
        ["完成时间", report.completed_at.strftime("%Y-%m-%d %H:%M:%S") if report.completed_at else "-"],
    ]

    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

    ws_summary.column_dimensions['A'].width = 15
    ws_summary.column_dimensions['B'].width = 25

    # Sheet 2: 用例明细
    ws_detail = wb.create_sheet("用例明细")
    headers = ["序号", "用例编号", "用例名称", "状态", "类型", "耗时(ms)", "备注"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws_detail.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for row_idx, s in enumerate(scenarios, 2):
        values = [
            row_idx - 1,
            s.case_code or "-",
            s.scenario_name,
            STATUS_LABELS.get(s.status, s.status),
            "自动" if s.execution_type == "automated" else "手动",
            s.duration_ms or "-",
            s.remark or "",
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws_detail.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            # 状态列着色
            if col_idx == 4 and s.status in STATUS_COLORS:
                cell.fill = PatternFill(start_color=STATUS_COLORS[s.status], end_color=STATUS_COLORS[s.status], fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal='center')

    # 设置列宽
    widths = [6, 18, 40, 10, 8, 12, 30]
    for i, w in enumerate(widths, 1):
        ws_detail.column_dimensions[chr(64 + i)].width = w

    # 输出
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
