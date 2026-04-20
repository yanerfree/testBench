import json
import uuid

from fastapi import APIRouter, Depends, Query, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from starlette.status import HTTP_201_CREATED

from app.core.exceptions import AppError
from app.deps.auth import require_project_role
from app.deps.db import get_db
from app.models.user import User
from app.schemas.case import BatchCaseRequest, CaseResponse, CopyFromBranchRequest, CreateCaseRequest, UpdateCaseRequest
from app.schemas.common import MessageResponse
from app.services import case_service, folder_service, import_service

router = APIRouter(prefix="/api/projects/{project_id}/branches/{branch_id}/cases", tags=["cases"])


@router.post("/import")
async def import_cases(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_project_role("project_admin", "developer", "tester")),
):
    """导入用例文件（支持 .json 和 .xlsx 格式）"""
    filename = file.filename or ""
    if not (filename.endswith(".json") or filename.endswith(".xlsx")):
        raise AppError(code="INVALID_FILE", message="仅接受 .json 或 .xlsx 文件", status_code=400)

    content = await file.read()
    if len(content) > 50 * 1024 * 1024:
        raise AppError(code="FILE_TOO_LARGE", message="文件大小不能超过 50MB", status_code=400)

    if filename.endswith(".json"):
        try:
            data = json.loads(content)
        except json.JSONDecodeError as e:
            raise AppError(
                code="JSON_PARSE_ERROR",
                message=f"JSON 解析失败：第 {e.lineno} 行",
                status_code=400,
                detail=str(e),
            )
        cases_list = data.get("cases", [])
        if not isinstance(cases_list, list):
            raise AppError(code="INVALID_FORMAT", message="JSON 中缺少 cases 数组", status_code=400)
    else:
        cases_list = _parse_excel_to_cases(content)

    summary = await import_service.import_cases(session, branch_id, cases_list)
    return {"data": summary}


def _parse_excel_to_cases(content: bytes) -> list[dict]:
    """解析导出的 Excel 文件为用例列表（兼容 export/excel 导出格式）。"""
    import io
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    headers = [str(h).strip() if h else "" for h in rows[0]]
    col = {h: i for i, h in enumerate(headers)}

    cases = []
    for row in rows[1:]:
        def get(name, default=""):
            idx = col.get(name)
            if idx is None or idx >= len(row) or row[idx] is None:
                return default
            return str(row[idx]).strip()

        title = get("标题")
        if not title:
            continue

        module = get("模块")
        submodule = get("子模块")
        case_type = get("测试类型", "api").lower()
        priority = get("优先级", "P2")
        tea_id = get("TEA ID") or get("用例ID") or f"excel-{uuid.uuid4().hex[:8]}"

        status_map = {"已自动化": "automated", "待自动化": "pending", "脚本已移除": "script_removed"}
        auto_status_raw = get("自动化状态", "pending")
        auto_status = status_map.get(auto_status_raw, auto_status_raw)

        script_file = get("脚本文件")
        script_func = get("脚本函数")
        script_ref = {}
        if script_file:
            script_ref["file"] = script_file
        if script_func:
            script_ref["func"] = script_func

        steps_text = get("测试步骤")
        steps = []
        if steps_text:
            for line in steps_text.split("\n"):
                line = line.strip()
                if line:
                    import re
                    line = re.sub(r"^\d+\.\s*", "", line)
                    steps.append({"action": line})

        cases.append({
            "tea_id": tea_id,
            "title": title,
            "type": case_type,
            "module": module,
            "submodule": submodule,
            "priority": priority,
            "script_ref": script_ref or None,
            "preconditions": get("前置条件") or None,
            "expected_result": get("预期结果") or None,
            "steps": steps or None,
            "remark": get("备注") or None,
        })

    wb.close()
    return cases


@router.post("", status_code=HTTP_201_CREATED)
async def create_case(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    body: CreateCaseRequest,
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_project_role("project_admin", "developer", "tester")),
):
    """手动创建用例"""
    case = await case_service.create_case(session, branch_id, body)
    return {
        "data": CaseResponse.model_validate(case, from_attributes=True).model_dump(by_alias=True)
    }


@router.get("")
async def list_cases(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=100, alias="pageSize"),
    case_type: str | None = Query(default=None, alias="type"),
    folder_id: uuid.UUID | None = Query(default=None, alias="folderId"),
    priority: str | None = Query(default=None),
    automation_status: str | None = Query(default=None, alias="automationStatus"),
    is_flaky: bool | None = Query(default=None, alias="isFlaky"),
    keyword: str | None = Query(default=None),
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_project_role("project_admin", "developer", "tester", "guest")),
):
    """用例列表（分页 + 多条件筛选）"""
    cases, total = await case_service.list_cases(
        session, branch_id, page, page_size,
        case_type=case_type, folder_id=folder_id, priority=priority,
        automation_status=automation_status, is_flaky=is_flaky, keyword=keyword,
    )
    return {
        "data": [
            CaseResponse.model_validate(c, from_attributes=True).model_dump(by_alias=True)
            for c in cases
        ],
        "pagination": {"page": page, "pageSize": page_size, "total": total},
    }


@router.get("/{case_id}")
async def get_case(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    case_id: uuid.UUID,
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_project_role("project_admin", "developer", "tester", "guest")),
):
    """用例详情"""
    case = await case_service.get_case(session, case_id)
    return {
        "data": CaseResponse.model_validate(case, from_attributes=True).model_dump(by_alias=True)
    }


@router.put("/{case_id}")
async def update_case(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    case_id: uuid.UUID,
    body: UpdateCaseRequest,
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_project_role("project_admin", "developer", "tester")),
):
    """更新用例"""
    case = await case_service.update_case(session, case_id, body)
    return {
        "data": CaseResponse.model_validate(case, from_attributes=True).model_dump(by_alias=True)
    }


@router.post("/batch")
async def batch_cases(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    body: BatchCaseRequest,
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_project_role("project_admin", "developer", "tester")),
):
    """批量操作用例（移动/归档/取消归档/修改优先级/标记Flaky）"""
    result = await case_service.batch_cases(
        session, branch_id,
        action=body.action,
        case_ids=body.case_ids,
        folder_id=body.folder_id,
        priority=body.priority,
    )
    return {"data": result}


@router.delete("/{case_id}")
async def delete_case(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    case_id: uuid.UUID,
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_project_role("project_admin", "developer", "tester")),
):
    """软删除用例（标记 deleted_at）"""
    await case_service.delete_case(session, case_id)
    return MessageResponse(message="删除成功").model_dump()


@router.post("/copy-from")
async def copy_from_branch(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    body: CopyFromBranchRequest,
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_project_role("project_admin", "developer", "tester")),
):
    """从其他分支复制用例到当前分支（深拷贝）"""
    result = await case_service.copy_cases_from_branch(
        session, branch_id, body.source_branch_id, body.case_ids
    )
    return {"data": result}


# ---- 用例目录 ----

folders_router = APIRouter(
    prefix="/api/projects/{project_id}/branches/{branch_id}/folders", tags=["folders"]
)


@folders_router.get("")
async def list_folders(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_project_role("project_admin", "developer", "tester", "guest")),
):
    """目录树（含用例计数）"""
    tree = await folder_service.list_folder_tree(session, branch_id)
    return {"data": tree}


@folders_router.post("", status_code=HTTP_201_CREATED)
async def create_folder(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    name: str = Query(..., min_length=1, max_length=100),
    parent_id: uuid.UUID | None = Query(default=None, alias="parentId"),
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_project_role("project_admin", "developer", "tester")),
):
    """创建模块/子模块目录"""
    folder = await folder_service.create_folder(session, branch_id, name, parent_id)
    return {"data": folder}


@folders_router.delete("/{folder_id}")
async def delete_folder(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    folder_id: uuid.UUID,
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_project_role("project_admin")),
):
    """删除目录（空目录才可删除）"""
    await folder_service.delete_folder(session, folder_id)
    return MessageResponse(message="删除成功").model_dump()


@router.get("/export/excel")
async def export_cases_excel(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_project_role("project_admin", "developer", "tester", "guest")),
    keyword: str | None = Query(default=None),
    automation_status: str | None = Query(default=None, alias="automationStatus"),
    folder_id: uuid.UUID | None = Query(default=None, alias="folderId"),
):
    """导出用例为 Excel"""
    import io
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    cases, _ = await case_service.list_cases(
        session, branch_id, page=1, page_size=10000,
        keyword=keyword, automation_status=automation_status, folder_id=folder_id,
    )

    # 加载目录映射: folder_id → (模块名, 子模块名)
    from sqlalchemy import select
    from app.models.case import CaseFolder
    folder_result = await session.execute(select(CaseFolder).where(CaseFolder.branch_id == branch_id))
    all_folders = {f.id: f for f in folder_result.scalars().all()}

    def get_folder_names(fid):
        if not fid or fid not in all_folders:
            return "", ""
        folder = all_folders[fid]
        if folder.parent_id and folder.parent_id in all_folders:
            return all_folders[folder.parent_id].name, folder.name
        return folder.name, ""

    wb = Workbook()
    ws = wb.active
    ws.title = "用例列表"

    headers = [
        "用例ID", "标题", "模块", "子模块", "测试类型", "优先级",
        "自动化状态", "来源", "Flaky",
        "前置条件", "测试步骤", "预期结果",
        "脚本文件", "脚本函数", "TEA ID",
        "备注", "创建时间", "更新时间",
    ]

    header_fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")
    header_font = Font(bold=True, size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    status_map = {"automated": "已自动化", "pending": "待自动化", "script_removed": "脚本已移除", "archived": "已归档"}

    for row_idx, c in enumerate(cases, 2):
        steps_text = ""
        if c.steps:
            steps_text = "\n".join(f"{i+1}. {s.get('action', s) if isinstance(s, dict) else s}" for i, s in enumerate(c.steps))

        module_name, sub_module_name = get_folder_names(c.folder_id)

        row = [
            c.case_code or "",
            c.title or "",
            module_name,
            sub_module_name,
            (c.type or "").upper(),
            c.priority or "",
            status_map.get(c.automation_status, c.automation_status or ""),
            "导入" if c.source == "imported" else "手动",
            "是" if c.is_flaky else "否",
            c.preconditions or "",
            steps_text,
            c.expected_result or "",
            c.script_ref_file or "",
            c.script_ref_func or "",
            c.tea_id or "",
            c.remark or "",
            c.created_at.strftime("%Y-%m-%d %H:%M") if c.created_at else "",
            c.updated_at.strftime("%Y-%m-%d %H:%M") if c.updated_at else "",
        ]
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    col_widths = [18, 40, 12, 12, 8, 6, 12, 6, 5, 30, 50, 30, 40, 25, 20, 20, 18, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=cases-export.xlsx"},
    )
